from math import floor

from django.db.models import Sum
from django.db.models.functions import Coalesce
from django.contrib import admin
from django.http import HttpResponse
from django.template.loader import get_template
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from django.contrib import messages

from django.conf import settings
from account.models import EmployeeSalary, SalarySheet, SalaryDisbursement
from config.utils.pdf import PDF


class FestivalBonusAction(admin.ModelAdmin):
    actions = (
        'export_bankasia_salary_acc_dis_excel',
        'export_salary_account_dis_excel',
        'export_bonus_account_dis_pdf',
    )

    @admin.action(description='Export Bank Asia Salary Account Disbursements (Excel)')
    def export_bankasia_salary_acc_dis_excel(self, request, queryset):
        salary_disbursement = SalaryDisbursement.objects.filter(disbursement_type='salary_account').first()
        return self.export_in_xl_bankasia(
            queryset, ('employee__in', salary_disbursement.employee.all()),
        )
    

    @admin.action(description='Export Salary Account Disbursements (Excel)')
    def export_salary_account_dis_excel(self, request, queryset):
        salary_disbursement = SalaryDisbursement.objects.filter(disbursement_type='salary_account').first()
        return self.export_in_xl_dbbl(
            queryset, ('employee__in', salary_disbursement.employee.all()),
        )
    
    @admin.action(description='Export Bonus Account Disbursements (PDF)')
    def export_bonus_account_dis_pdf(self, request, queryset):
        return self.bonus_pdf(
            queryset=queryset,
            filter=('disbursement_type', 'salary_account'),
            bank={
                'ref': 'Mediuswareltd',
                'account_name': 'Mediusware Ltd.',
                'account_number': '1481100038741'
            },
        )
    
    def bonus_pdf(self, queryset, filter=None, bank=None):
        salary_disbursement = SalaryDisbursement.objects.filter(filter).first()

        festival_bonus_sheet = queryset.first()
        employee_festival_bonus_set = festival_bonus_sheet.employeefestivalbonus_set.filter(
            employee__in=salary_disbursement.employee.all().values_list("id", flat=True)
        )

        pdf = PDF()
        pdf.context = {
            'festival_bonus_sheet':festival_bonus_sheet,
            'employee_festival_bonus_set': employee_festival_bonus_set,
            'bank': bank,
            'seal': f"{settings.STATIC_ROOT}/stationary/sign_md.png"
        }
        pdf.template_path = 'letters/bonus_pdf_v2.html'
        return pdf.render_to_pdf()
    

    def export_in_xl_bankasia(self, queryset, query_filter=None):
        wb = Workbook()
        work_sheets = {}
        for festival_bonus_sheet in queryset:

            festival_bonus_sheet.total_value = 0

            work_sheet = wb.create_sheet(title=str(festival_bonus_sheet.date))

            work_sheet.append([
                'Employee Name', 
                'Account Number', 
                'Dr./Cr.',  
                'Transaction Amount', 
                'Chqser',
                'Chqnum',
                'Chqdat',
                'Remarks',
            ])

            work_sheet.append([
                settings.COMPANY_ACCOUNT_NAME,
                settings.COMPANY_ACCOUNT_NO,
                'D',
                '0',
                '',
                '',
                '',
                f'Festival Bonus for {festival_bonus_sheet.date.strftime("%b, %Y")}',
            ])

            employee_festival_bonuses = festival_bonus_sheet.employeefestivalbonus_set
            if query_filter is not None:
                employee_festival_bonuses = festival_bonus_sheet.employeefestivalbonus_set.filter(query_filter).all()
            
            for employee_festival_bonus in employee_festival_bonuses.all():

                festival_bonus_sheet.total_value += floor(employee_festival_bonus.amount)
                bank_account = employee_festival_bonus.employee.bankaccount_set.filter(default=True, is_approved=True).last()
                
                work_sheet.append([
                    employee_festival_bonus.employee.full_name,
                    bank_account.account_number if bank_account else '-',
                    "C",
                    floor(employee_festival_bonus.amount),
                    "", "", "",
                    f'Festival Bonus for {festival_bonus_sheet.date.strftime("%b, %Y")}',
                ])
            
            work_sheet["D2"] = festival_bonus_sheet.total_value
            work_sheets[str(festival_bonus_sheet.id)] = work_sheet
        wb.remove(wb['Sheet'])
        response = HttpResponse(content=save_virtual_workbook(wb), content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename=FestivalBonusSheet.xlsx'
        return response


    def export_in_xl_dbbl(self, queryset, query_filter=None):
        """

        @param queryset:
        @param query_filter:
        @return:
        """
        wb = Workbook()
        work_sheets = {}
        for festival_bonus_sheet in queryset:

            festival_bonus_sheet.total_value = 0

            work_sheet = wb.create_sheet(title=str(festival_bonus_sheet.date))

            work_sheet.append(['name', 'Basic Salary', 'Bonus Amount',  'Bank Name', 'Bank Number'])

            employee_festival_bonuses = festival_bonus_sheet.employeefestivalbonus_set
            if query_filter is not None:
                employee_festival_bonuses = festival_bonus_sheet.employeefestivalbonus_set.filter(query_filter).all()
            
            for employee_festival_bonus in employee_festival_bonuses.all():
                festival_bonus_sheet.total_value += floor(employee_festival_bonus.amount)
                bank_account = employee_festival_bonus.employee.bankaccount_set.filter(default=True, is_approved=True).last()

                salary_history = employee_festival_bonus.employee.salaryhistory_set.filter(
                    active_from__lte=festival_bonus_sheet.date.replace(day=1)
                ).last()

                basic_salary = 0
                if salary_history:
                    basic_salary = (salary_history.payable_salary / 100) * employee_festival_bonus.employee.pay_scale.basic

                work_sheet.append([
                    employee_festival_bonus.employee.full_name,
                    basic_salary,
                    employee_festival_bonus.amount,
                    bank_account.bank.name if bank_account else '',
                    bank_account.account_number if bank_account else ''
                ])
            work_sheet.append(['', 'Total', festival_bonus_sheet.total_value])
            work_sheets[str(festival_bonus_sheet.id)] = work_sheet
        wb.remove(wb['Sheet'])
        response = HttpResponse(content=save_virtual_workbook(wb), content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename=FestivalBonusSheet.xlsx'
        return response

